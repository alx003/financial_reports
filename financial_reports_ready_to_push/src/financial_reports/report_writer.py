"""Render extracted line-item series into a broker-report-style workbook:
quarters grouped under each fiscal year with a spacer + FY total column,
a dark header band, indented sub-items, and blank spacer rows between the
three statements -- styled after the Langenberg & Co. EPS Model layout.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .metric_map import BALANCE_SHEET, CASH_FLOW, INCOME_STATEMENT, LineItem
from .period_extraction import PeriodKey

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16)
SECTION_FONT = Font(name="Arial", bold=True, size=12)
SUBTOTAL_FONT = Font(name="Arial", bold=True, size=10)
NORMAL_FONT = Font(name="Arial", size=10)
LABEL_COL_WIDTH = 38
DATA_COL_WIDTH = 12
NUM_FORMAT = '#,##0;(#,##0);"-"'
EPS_FORMAT = '$#,##0.00;($#,##0.00);"-"'
THIN_TOP = Border(top=Side(style="thin"))
DOUBLE_TOP = Border(top=Side(style="double"))


def _build_columns(periods: list[PeriodKey]) -> list[tuple[str, PeriodKey | None, bool]]:
    """Group periods by fiscal year: Q1 Q2 Q3 Q4 [spacer] FY [spacer-col].
    Returns list of (column_label, period_or_None, is_fy_total)."""
    by_fy: dict[int, list[PeriodKey]] = {}
    for pk in periods:
        by_fy.setdefault(pk.fiscal_year, []).append(pk)

    columns: list[tuple[str, PeriodKey | None, bool]] = []
    for fy in sorted(by_fy):
        quarters = sorted([p for p in by_fy[fy] if p.fiscal_period.startswith("Q")],
                           key=lambda p: p.fiscal_period)
        fy_total = next((p for p in by_fy[fy] if p.fiscal_period == "FY"), None)
        for q in quarters:
            columns.append((q.fiscal_period, q, False))
        if fy_total is not None:
            columns.append((str(fy), fy_total, True))
        columns.append(("", None, False))  # spacer column between fiscal years
    return columns


def _is_eps_or_shares(label: str) -> bool:
    return "EPS" in label or "Shares" in label


def _write_statement(ws, start_row: int, title: str, items: tuple[LineItem, ...],
                      all_series: dict[str, dict[PeriodKey, float]],
                      columns: list[tuple[str, PeriodKey | None, bool]]) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1

    fy_header_row = row
    quarter_header_row = row + 1
    col = 2
    fy_label_written: dict[int, int] = {}
    for label, pk, is_total in columns:
        if pk is not None:
            ws.cell(row=quarter_header_row, column=col, value=label)
            if is_total or pk.fiscal_period == "FY":
                if pk.fiscal_year not in fy_label_written:
                    ws.cell(row=fy_header_row, column=col, value=pk.fiscal_year)
                    fy_label_written[pk.fiscal_year] = col
        col += 1
    for c in range(2, col):
        for r in (fy_header_row, quarter_header_row):
            cell = ws.cell(row=r, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
    ws.cell(row=fy_header_row, column=1).fill = HEADER_FILL
    ws.cell(row=quarter_header_row, column=1).fill = HEADER_FILL
    row = quarter_header_row + 1

    for item in items:
        series = all_series.get(item.label, {})
        label_cell = ws.cell(row=row, column=1, value=("  " * item.indent) + item.label)
        label_cell.font = SUBTOTAL_FONT if item.is_subtotal else NORMAL_FONT
        col = 2
        for _, pk, _ in columns:
            if pk is not None and pk in series:
                cell = ws.cell(row=row, column=col, value=series[pk])
                cell.number_format = EPS_FORMAT if _is_eps_or_shares(item.label) else NUM_FORMAT
                cell.font = SUBTOTAL_FONT if item.is_subtotal else NORMAL_FONT
                if item.is_subtotal:
                    cell.border = THIN_TOP
            col += 1
        if item.is_subtotal:
            label_cell.border = THIN_TOP
        row += 1

    return row + 2  # blank spacer rows before next statement


def write_workbook(all_series: dict[str, dict[PeriodKey, float]], periods: list[PeriodKey],
                    ticker: str, company_name: str, out_path: Path, units_note: str = "USD, as reported") -> None:
    columns = _build_columns(periods)
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Statements"

    ws.cell(row=1, column=1, value=f"{company_name} ({ticker})").font = TITLE_FONT
    ws.cell(row=2, column=1, value="STANDARD FINANCIAL STATEMENTS").font = SECTION_FONT
    ws.cell(row=3, column=1, value=f"Units: {units_note}  |  Source: SEC EDGAR XBRL (us-gaap), standard filings only").font = NORMAL_FONT

    row = 5
    row = _write_statement(ws, row, "INCOME STATEMENT", INCOME_STATEMENT, all_series, columns)
    row = _write_statement(ws, row, "BALANCE SHEET", BALANCE_SHEET, all_series, columns)
    row = _write_statement(ws, row, "CASH FLOW STATEMENT", CASH_FLOW, all_series, columns)

    ws.column_dimensions["A"].width = LABEL_COL_WIDTH
    for c in range(2, len(columns) + 2):
        ws.column_dimensions[get_column_letter(c)].width = DATA_COL_WIDTH
    ws.freeze_panes = "B6"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_csv(all_series: dict[str, dict[PeriodKey, float]], periods: list[PeriodKey],
              out_path: Path) -> None:
    """Long-form CSV (one row per line item x period) for easy ingestion
    into pandas/BI tools, mirroring what export_git_dataset.py does in the
    sibling us-company-sec-dataset repo."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    statement_for_label = {}
    for items, name in ((INCOME_STATEMENT, "income_statement"),
                         (BALANCE_SHEET, "balance_sheet"),
                         (CASH_FLOW, "cash_flow")):
        for item in items:
            statement_for_label[item.label] = (name, item.is_subtotal, item.indent)

    with out_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["statement", "line_item", "fiscal_year", "fiscal_period",
                          "period_end_date", "value", "is_subtotal"])
        for label, series in all_series.items():
            statement, is_subtotal, _ = statement_for_label.get(label, ("", False, 0))
            for pk, val in sorted(series.items(), key=lambda kv: kv[0].sort_key()):
                writer.writerow([statement, label, pk.fiscal_year, pk.fiscal_period,
                                  pk.end_date.isoformat(), val, is_subtotal])
