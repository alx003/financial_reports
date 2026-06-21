"""Render extracted line-item series into a broker-report-style workbook:
quarters grouped under each fiscal year with a spacer + FY total column,
a dark header band, indented sub-items, and blank spacer rows between the
three statements -- styled after the Langenberg & Co. EPS Model layout.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .dashboard_metrics import DASHBOARD_METRICS, DASHBOARD_SECTIONS, DashboardMetric
from .metric_map import BALANCE_SHEET, CASH_FLOW, INCOME_STATEMENT, LineItem
from .period_extraction import PeriodKey

HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(name="Arial", color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=16)
SECTION_FONT = Font(name="Arial", bold=True, size=12)
SUBTOTAL_FONT = Font(name="Arial", bold=True, size=10)
NORMAL_FONT = Font(name="Arial", size=10)
LABEL_COL_WIDTH = 38
DATA_COL_WIDTH = 12
NUM_FORMAT = '#,##0;(#,##0);"-"'
EPS_FORMAT = '$#,##0.00;($#,##0.00);"-"'
PCT_FORMAT = '0.0%;(0.0%);"-"'
THIN_TOP = Border(top=Side(style="thin"))
DOUBLE_TOP = Border(top=Side(style="double"))
SECTION_SUBHEADER_FILL = PatternFill("solid", fgColor="EDEDED")
SECTION_SUBHEADER_FONT = Font(name="Arial Narrow", bold=True, size=10)
ITALIC_NOTE_FONT = Font(name="Arial", italic=True, size=9)
DASHBOARD_TREND_QUARTERS = 16  # matches the BA model's "trend analysis of last 16 quarters"
OPEN_FUTURE_QUARTERS = 4       # extra blank quarter columns left open to the right for future fills


def _build_columns(periods: list[PeriodKey]) -> list[tuple[str, PeriodKey | None, bool]]:
    """Group periods by fiscal year: Q1 Q2 Q3 Q4 [spacer] FY [spacer-col].
    Returns list of (column_label, period_or_None, is_fy_total)."""
    by_fy: dict[int, list[PeriodKey]] = {}
    for pk in periods:
        by_fy.setdefault(pk.fiscal_year, []).append(pk)

    columns: list[tuple[str, PeriodKey | None, bool]] = []
    for fy in sorted(by_fy):
        quarters = sorted([p for p in by_fy[fy] if p.fiscal_period.startswith("Q")],
                           key=lambda p: p.fiscal_period)
        fy_total = next((p for p in by_fy[fy] if p.fiscal_period == "FY"), None)
        for q in quarters:
            columns.append((q.fiscal_period, q, False))
        if fy_total is not None:
            columns.append((str(fy), fy_total, True))
        columns.append(("", None, False))  # spacer column between fiscal years
    return columns


def _is_eps_or_shares(label: str) -> bool:
    return "EPS" in label or "Shares" in label


def _collect_quarterly_periods(periods: list[PeriodKey], count: int) -> list[PeriodKey]:
    """Return the most recent `count` quarter-only periods (Q1-Q4, no FY
    totals), oldest-to-newest, for the trend-chart window."""
    quarters = [p for p in periods if p.fiscal_period.startswith("Q")]
    quarters.sort(key=lambda p: p.sort_key())
    return quarters[-count:] if len(quarters) > count else quarters


def write_dashboard(ws, start_row: int, all_series: dict[str, dict[PeriodKey, float]],
                     periods: list[PeriodKey]) -> int:
    """Write the SNAPSHOT/DASHBOARD trend-chart block: a 16-quarter rolling
    window per metric, followed by deliberately blank columns for
    not-yet-reported future quarters, a spacer, then live L10/L5/L3
    rolling-average formulas -- mirroring the Langenberg & Co. BA model's
    Dashboard layout and mechanics so the same column-extension approach
    (see add_quarter_column.py in the BA model) applies here too.

    The future-quarter columns are intentionally left blank rather than
    populated with placeholders: when the next 10-Q lands, those columns
    get filled in (manually, or by a future automated step) and the
    L10/L5/L3 formulas -- already anchored with one quarter of headroom --
    pick the new data up automatically.
    """
    row = start_row
    ws.cell(row=row, column=1, value="DASHBOARD").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1,
             value="Trend analysis of last 16 quarters. Columns to the right of the "
                   "latest reported quarter are intentionally left blank for future updates.").font = ITALIC_NOTE_FONT
    row += 2

    trend_quarters = _collect_quarterly_periods(periods, DASHBOARD_TREND_QUARTERS)
    if not trend_quarters:
        ws.cell(row=row, column=1, value="(no quarterly data available yet)").font = NORMAL_FONT
        return row + 2

    n_qtr_cols = len(trend_quarters)
    open_start_col = 2 + n_qtr_cols                       # first blank "future quarter" column
    open_end_col = open_start_col + OPEN_FUTURE_QUARTERS - 1
    spacer_col = open_end_col + 1
    l10_col, l5_col, l3_col = spacer_col + 1, spacer_col + 2, spacer_col + 3

    # --- Header row: quarter labels, open blanks, then L10/L5/L3 ---
    header_row = row
    for i, pk in enumerate(trend_quarters):
        c = 2 + i
        cell = ws.cell(row=header_row, column=c, value=pk.label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for c in range(open_start_col, open_end_col + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL  # styled but left empty -- future quarters
    for c, label in ((l10_col, "L10"), (l5_col, "L5"), (l3_col, "L3")):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.cell(row=header_row, column=1).fill = HEADER_FILL
    row = header_row + 1

    first_qtr_col = 2  # leftmost quarter column (column B)

    for section in DASHBOARD_SECTIONS:
        section_metrics = [m for m in DASHBOARD_METRICS if m.section == section]
        if not section_metrics:
            continue
        sub_cell = ws.cell(row=row, column=1, value=section)
        sub_cell.font = SECTION_SUBHEADER_FONT
        sub_cell.fill = SECTION_SUBHEADER_FILL
        for c in range(2, l3_col + 1):
            ws.cell(row=row, column=c).fill = SECTION_SUBHEADER_FILL
        row += 1

        for metric in section_metrics:
            label_cell = ws.cell(row=row, column=1, value=("  " * metric.indent) + metric.label)
            label_cell.font = NORMAL_FONT
            fmt = PCT_FORMAT if metric.is_percent else NUM_FORMAT

            row_last_qtr_col = first_qtr_col - 1  # tracks this row's own last populated quarter
            for i, pk in enumerate(trend_quarters):
                val = metric.func(all_series, pk)
                if val is not None:
                    c = 2 + i
                    cell = ws.cell(row=row, column=c, value=val)
                    cell.number_format = fmt
                    cell.font = NORMAL_FONT
                    row_last_qtr_col = c
            # Future-quarter columns: left genuinely blank (no value, no
            # formula) -- this is the "open" right side for new data.

            # Live rolling-average formulas, anchored to THIS ROW's own
            # last populated quarter column (not necessarily the same as
            # other rows, since different metrics can have different data
            # availability -- e.g. one line item reported a quarter later
            # than another). If a row has no data at all, the L10/L5/L3
            # cells are left blank too rather than producing a formula
            # that references nothing.
            if row_last_qtr_col >= first_qtr_col:
                last_col_letter = get_column_letter(row_last_qtr_col)
                l10_start = get_column_letter(max(first_qtr_col, row_last_qtr_col - 9))
                l5_start = get_column_letter(max(first_qtr_col, row_last_qtr_col - 4))
                l3_start = get_column_letter(max(first_qtr_col, row_last_qtr_col - 2))
                for col, start_letter in ((l10_col, l10_start), (l5_col, l5_start), (l3_col, l3_start)):
                    formula = f'=IFERROR(AVERAGE({start_letter}{row}:{last_col_letter}{row}),"")'
                    cell = ws.cell(row=row, column=col, value=formula)
                    cell.number_format = fmt
                    cell.font = NORMAL_FONT
            row += 1
        row += 1  # blank spacer row between sub-sections

    return row + 1


def _write_statement(ws, start_row: int, title: str, items: tuple[LineItem, ...],
                      all_series: dict[str, dict[PeriodKey, float]],
                      columns: list[tuple[str, PeriodKey | None, bool]]) -> int:
    row = start_row
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    row += 1

    fy_header_row = row
    quarter_header_row = row + 1
    col = 2
    fy_label_written: dict[int, int] = {}
    for label, pk, is_total in columns:
        if pk is not None:
            ws.cell(row=quarter_header_row, column=col, value=label)
            if is_total or pk.fiscal_period == "FY":
                if pk.fiscal_year not in fy_label_written:
                    ws.cell(row=fy_header_row, column=col, value=pk.fiscal_year)
                    fy_label_written[pk.fiscal_year] = col
        col += 1
    for c in range(2, col):
        for r in (fy_header_row, quarter_header_row):
            cell = ws.cell(row=r, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
    ws.cell(row=fy_header_row, column=1).fill = HEADER_FILL
    ws.cell(row=quarter_header_row, column=1).fill = HEADER_FILL
    row = quarter_header_row + 1

    for item in items:
        series = all_series.get(item.label, {})
        label_cell = ws.cell(row=row, column=1, value=("  " * item.indent) + item.label)
        label_cell.font = SUBTOTAL_FONT if item.is_subtotal else NORMAL_FONT
        col = 2
        for _, pk, _ in columns:
            if pk is not None and pk in series:
                cell = ws.cell(row=row, column=col, value=series[pk])
                cell.number_format = EPS_FORMAT if _is_eps_or_shares(item.label) else NUM_FORMAT
                cell.font = SUBTOTAL_FONT if item.is_subtotal else NORMAL_FONT
                if item.is_subtotal:
                    cell.border = THIN_TOP
            col += 1
        if item.is_subtotal:
            label_cell.border = THIN_TOP
        row += 1

    return row + 2  # blank spacer rows before next statement


def write_workbook(all_series: dict[str, dict[PeriodKey, float]], periods: list[PeriodKey],
                    ticker: str, company_name: str, out_path: Path, units_note: str = "USD, as reported") -> None:
    columns = _build_columns(periods)
    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Statements"

    ws.cell(row=1, column=1, value=f"{company_name} ({ticker})").font = TITLE_FONT
    ws.cell(row=2, column=1, value="STANDARD FINANCIAL STATEMENTS").font = SECTION_FONT
    ws.cell(row=3, column=1, value=f"Units: {units_note}  |  Source: SEC EDGAR XBRL (us-gaap), standard filings only").font = NORMAL_FONT

    row = 5
    row = write_dashboard(ws, row, all_series, periods)
    row = _write_statement(ws, row, "INCOME STATEMENT", INCOME_STATEMENT, all_series, columns)
    row = _write_statement(ws, row, "BALANCE SHEET", BALANCE_SHEET, all_series, columns)
    row = _write_statement(ws, row, "CASH FLOW STATEMENT", CASH_FLOW, all_series, columns)

    ws.column_dimensions["A"].width = LABEL_COL_WIDTH
    max_col = max(len(columns) + 2, 2 + DASHBOARD_TREND_QUARTERS + OPEN_FUTURE_QUARTERS + 4)
    for c in range(2, max_col):
        ws.column_dimensions[get_column_letter(c)].width = DATA_COL_WIDTH
    ws.freeze_panes = "B6"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_csv(all_series: dict[str, dict[PeriodKey, float]], periods: list[PeriodKey],
              out_path: Path) -> None:
    """Long-form CSV (one row per line item x period) for easy ingestion
    into pandas/BI tools, mirroring what export_git_dataset.py does in the
    sibling us-company-sec-dataset repo."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    statement_for_label = {}
    for items, name in ((INCOME_STATEMENT, "income_statement"),
                         (BALANCE_SHEET, "balance_sheet"),
                         (CASH_FLOW, "cash_flow")):
        for item in items:
            statement_for_label[item.label] = (name, item.is_subtotal, item.indent)

    with out_path.open("w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["statement", "line_item", "fiscal_year", "fiscal_period",
                          "period_end_date", "value", "is_subtotal"])
        for label, series in all_series.items():
            statement, is_subtotal, _ = statement_for_label.get(label, ("", False, 0))
            for pk, val in sorted(series.items(), key=lambda kv: kv[0].sort_key()):
                writer.writerow([statement, label, pk.fiscal_year, pk.fiscal_period,
                                  pk.end_date.isoformat(), val, is_subtotal])
